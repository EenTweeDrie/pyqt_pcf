import pandas as pd
from PyQt6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QTableWidget,
                             QTableWidgetItem, QLabel, QComboBox, QPushButton,
                             QMessageBox, QHeaderView, QScrollArea, QFrame)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QFont
import openpyxl


class XlsxViewerWidget(QWidget):
    """Виджет для просмотра содержимого Excel файлов"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.current_file = None
        self.workbook = None
        self.init_ui()

    def init_ui(self):
        """Инициализация интерфейса"""
        layout = QVBoxLayout()

        # Применяем общий стиль приложения
        self.setStyleSheet("""
            QWidget {
                background-color: #53535A;
                color: #CCCEDB;
            }
            QLabel {
                background-color: #53535A;
                color: #CCCEDB;
                border: none;
            }
            QPushButton {
                background-color: #5D5D64;
                color: #CCCEDB;
                border: 1px solid #66666E;
                border-radius: 3px;
                padding: 5px 10px;
                min-height: 20px;
            }
            QPushButton:hover {
                background-color: #66666E;
                border: 1px solid #7A7A81;
            }
            QPushButton:pressed {
                background-color: #4A4A51;
                border: 1px solid #5D5D64;
            }
            QComboBox {
                background-color: #5D5D64;
                color: #CCCEDB;
                border: 1px solid #66666E;
                border-radius: 3px;
                padding: 5px;
                min-height: 20px;
            }
            QComboBox:hover {
                background-color: #66666E;
                border: 1px solid #7A7A81;
            }
            QComboBox::drop-down {
                border: none;
                background-color: #5D5D64;
            }
            QComboBox::down-arrow {
                image: none;
                border-left: 5px solid transparent;
                border-right: 5px solid transparent;
                border-top: 5px solid #CCCEDB;
                margin-right: 5px;
            }
            QTableWidget {
                background-color: #53535A;
                color: #CCCEDB;
                gridline-color: #5D5D64;
                border: 2px solid #5D5D64;
                selection-background-color: #1E88E5;
                alternate-background-color: #5D5D64;
            }
            QTableWidget::item {
                background-color: #53535A;
                color: #CCCEDB;
                border: none;
                padding: 5px;
            }
            QTableWidget::item:alternate {
                background-color: #5D5D64;
            }
            QTableWidget::item:selected {
                background-color: #1E88E5;
                color: white;
            }
            QHeaderView::section {
                background-color: #5D5D64;
                color: #CCCEDB;
                border: 1px solid #66666E;
                padding: 5px;
                font-weight: bold;
            }
            QScrollBar:vertical {
                background: #53535A;
                width: 15px;
                border-radius: 7px;
                margin: 0px;
            }
            QScrollBar::handle:vertical {
                background: #5D5D64;
                border-radius: 7px;
                min-height: 20px;
            }
            QScrollBar::handle:vertical:hover {
                background: #66666E;
            }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                height: 0px;
            }
            QScrollBar:horizontal {
                background: #53535A;
                height: 15px;
                border-radius: 7px;
                margin: 0px;
            }
            QScrollBar::handle:horizontal {
                background: #5D5D64;
                border-radius: 7px;
                min-width: 20px;
            }
            QScrollBar::handle:horizontal:hover {
                background: #66666E;
            }
            QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {
                width: 0px;
            }
        """)

        # Панель управления с заголовком и кнопками
        control_layout = QHBoxLayout()

        # Заголовок слева
        self.title_label = QLabel("Просмотр Excel файла")
        self.title_label.setFont(QFont("Arial", 12, QFont.Weight.Bold))
        self.title_label.setAlignment(Qt.AlignmentFlag.AlignLeft)
        control_layout.addWidget(self.title_label)

        # Добавляем растяжение для разделения заголовка и кнопок
        control_layout.addStretch()

        # Информационная панель (справа)
        self.info_label = QLabel("")
        self.info_label.setStyleSheet("color: #888; font-size: 12px;")
        control_layout.addWidget(self.info_label)

        # Выбор листа (справа)
        control_layout.addWidget(QLabel("Лист:"))
        self.sheet_combo = QComboBox()
        self.sheet_combo.currentTextChanged.connect(self.load_sheet)
        control_layout.addWidget(self.sheet_combo)

        # Кнопка обновления (справа)
        self.refresh_button = QPushButton("Обновить")
        self.refresh_button.clicked.connect(self.refresh_data)
        control_layout.addWidget(self.refresh_button)

        layout.addLayout(control_layout)

        # Информационная панель

        # Таблица для отображения данных
        self.table = QTableWidget()
        self.table.setAlternatingRowColors(True)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        layout.addWidget(self.table)

        self.setLayout(layout)

    def load_file(self, file_path):
        """Загрузка Excel файла"""
        try:
            self.current_file = file_path
            self.workbook = openpyxl.load_workbook(file_path, read_only=True)

            # Обновляем заголовок
            file_name = file_path.split('/')[-1].split('\\')[-1]
            self.title_label.setText(f"{file_name}")

            # Заполняем список листов
            self.sheet_combo.clear()
            self.sheet_combo.addItems(self.workbook.sheetnames)

            # Загружаем первый лист
            if self.workbook.sheetnames:
                self.load_sheet(self.workbook.sheetnames[0])

        except Exception as e:
            QMessageBox.warning(self, "Ошибка", f"Не удалось загрузить файл:\n{str(e)}")
            self.clear_data()

    def load_sheet(self, sheet_name):
        """Загрузка данных конкретного листа"""
        if not self.workbook or not sheet_name:
            return

        try:
            # Используем pandas для более удобного чтения
            df = pd.read_excel(self.current_file, sheet_name=sheet_name, engine='openpyxl')

            # Обновляем информацию
            rows, cols = df.shape
            self.info_label.setText(f"Строк: {rows}, Столбцов: {cols}")

            # Заполняем таблицу
            self.table.setRowCount(rows)
            self.table.setColumnCount(cols)

            # Устанавливаем заголовки столбцов
            self.table.setHorizontalHeaderLabels([str(col) for col in df.columns])

            # Заполняем данные
            for i in range(rows):
                for j in range(cols):
                    value = df.iloc[i, j]
                    if pd.isna(value):
                        item_text = ""
                    else:
                        item_text = str(value)

                    item = QTableWidgetItem(item_text)
                    item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)  # Только для чтения
                    self.table.setItem(i, j, item)

            # Автоподбор размера столбцов
            self.table.resizeColumnsToContents()

        except Exception as e:
            QMessageBox.warning(self, "Ошибка", f"Не удалось загрузить лист '{sheet_name}':\n{str(e)}")
            self.clear_data()

    def refresh_data(self):
        """Обновление данных"""
        if self.current_file:
            current_sheet = self.sheet_combo.currentText()
            self.load_file(self.current_file)
            # Восстанавливаем выбранный лист
            if current_sheet:
                index = self.sheet_combo.findText(current_sheet)
                if index >= 0:
                    self.sheet_combo.setCurrentIndex(index)

    def clear_data(self):
        """Очистка данных"""
        self.table.clear()
        self.table.setRowCount(0)
        self.table.setColumnCount(0)
        self.sheet_combo.clear()
        self.info_label.setText("")
        self.title_label.setText("Просмотр Excel файла")
        self.current_file = None
        if self.workbook:
            self.workbook.close()
            self.workbook = None
